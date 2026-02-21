"""
debug_logger.py
Writes a timestamped debug Excel file to Log_file/ showing every
intermediate value for each game — for comparing against Google Sheets.
"""

import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

LOG_DIR = r"C:\Users\chris\OneDrive\Documents\cbb_model\Log_file"

# ── Colors ────────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E5090"
LIGHT_BLUE  = "D6E4F0"
GOLD        = "F0B429"
LIGHT_GOLD  = "FFF3CD"
GREEN       = "27A148"
LIGHT_GREEN = "D4EDDA"
RED         = "C0392B"
LIGHT_RED   = "FDECEA"
GRAY        = "F5F5F5"
WHITE       = "FFFFFF"


def _cell_style(ws, row, col, value, bold=False, bg=None, font_color="000000",
                align="left", num_fmt=None, border=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=font_color, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    if num_fmt:
        c.number_format = num_fmt
    if border:
        thin = Side(style="thin", color="CCCCCC")
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c


def _section_header(ws, row, col, text, width_cols=4):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c.fill = PatternFill("solid", start_color=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    if width_cols > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + width_cols - 1)
    return c


def _row_pair(ws, row, label, t1_val, t2_val, t1_name, t2_name,
              num_fmt="#,##0.0000", highlight=False):
    bg = LIGHT_GOLD if highlight else None
    _cell_style(ws, row, 1, label, bg=bg, border=True)
    _cell_style(ws, row, 2, t1_val, bg=LIGHT_BLUE if not highlight else LIGHT_GOLD,
                align="right", num_fmt=num_fmt, border=True)
    _cell_style(ws, row, 3, t2_val, bg=LIGHT_GREEN if not highlight else LIGHT_GOLD,
                align="right", num_fmt=num_fmt, border=True)


def write_debug_excel(game_results: list[dict], run_date: str):
    """
    game_results: list of dicts, each containing the full debug info for one game.
    run_date: date string like '2026-02-20'
    """
    os.makedirs(LOG_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"debug_{run_date}_{timestamp}.xlsx"
    filepath = os.path.join(LOG_DIR, filename)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 16
    ws_sum.column_dimensions["C"].width = 16
    ws_sum.column_dimensions["D"].width = 14
    ws_sum.column_dimensions["E"].width = 14
    ws_sum.column_dimensions["F"].width = 14
    ws_sum.column_dimensions["G"].width = 14
    ws_sum.column_dimensions["H"].width = 12
    ws_sum.column_dimensions["I"].width = 12
    ws_sum.column_dimensions["J"].width = 14
    ws_sum.row_dimensions[1].height = 30
    ws_sum.row_dimensions[2].height = 22

    # Header
    _section_header(ws_sum, 1, 1, f"🏀 CBB Model Debug Log — {run_date}", width_cols=7)
    headers = ["Matchup", "Your Home", "Your Away", "KP Home", "KP Away", "Your Total", "Vegas Spread", "Vegas Total", "Edge Score", "Book"]
    for i, h in enumerate(headers, 1):
        c = ws_sum.cell(row=2, column=i, value=h)
        c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", start_color=MID_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, g in enumerate(game_results, 3):
        ws_sum.row_dimensions[i].height = 18
        matchup = f"{g['team1']} vs {g['team2']}"
        _cell_style(ws_sum, i, 1, matchup, border=True, bg=GRAY if i % 2 == 0 else WHITE)
        _cell_style(ws_sum, i, 2, g["team1_score"], align="center", num_fmt="0.0", border=True,
                    bg=LIGHT_BLUE)
        _cell_style(ws_sum, i, 3, g["team2_score"], align="center", num_fmt="0.0", border=True,
                    bg=LIGHT_BLUE)
        _cell_style(ws_sum, i, 4, g.get("kp_home_score"), align="center", num_fmt="0.0", border=True,
                    bg=LIGHT_GREEN)
        _cell_style(ws_sum, i, 5, g.get("kp_away_score"), align="center", num_fmt="0.0", border=True,
                    bg=LIGHT_GREEN)
        _cell_style(ws_sum, i, 6, g["total"], align="center", num_fmt="0.0", border=True)
        _cell_style(ws_sum, i, 7, g.get("kp_tempo"), align="center", num_fmt="0.0", border=True)
        _cell_style(ws_sum, i, 8, g.get("vegas_spread"), align="center", num_fmt="+0.0;-0.0", border=True,
                    bg=LIGHT_RED)
        _cell_style(ws_sum, i, 9, g.get("vegas_total"), align="center", num_fmt="0.0", border=True,
                    bg=LIGHT_RED)
        edge = g.get("edge_score")
        edge_bg = "FFD700" if edge and edge > 0.08 else (LIGHT_GREEN if edge and edge > 0.05 else None)
        _cell_style(ws_sum, i, 10, edge, align="center", num_fmt="0.0000", border=True, bg=edge_bg)

    # ── One detail sheet per game ─────────────────────────────────────────────
    for g in game_results:
        sheet_name = f"{g['team1'][:12]} v {g['team2'][:12]}"
        ws = wb.create_sheet(sheet_name)
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 20

        t1 = g["team1"]
        t2 = g["team2"]

        # Title row
        _section_header(ws, 1, 1, f"{t1} (Home)  vs  {t2} (Away)", width_cols=3)

        # Column headers
        for col, label in [(1, "Metric"), (2, t1), (3, t2)]:
            c = ws.cell(row=2, column=col, value=label)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = PatternFill("solid", start_color=MID_BLUE if col == 1 else
                                 (DARK_BLUE if col == 2 else "1a4a2a"))
            c.alignment = Alignment(horizontal="center")

        r = 3

        # ── FINAL SCORES (highlighted) ────────────────────────────────────────
        _section_header(ws, r, 1, "FINAL PROJECTIONS", width_cols=3); r += 1
        _row_pair(ws, r, "Projected Score", g["team1_score"], g["team2_score"], t1, t2,
                  num_fmt="0.0", highlight=True); r += 1
        _row_pair(ws, r, "KenPom Predicted Score", g.get("kp_home_score"), g.get("kp_away_score"),
                  t1, t2, num_fmt="0.0", highlight=True); r += 1
        _row_pair(ws, r, "Spread (+ = team favored)", g["spread"], -g["spread"],
                  t1, t2, num_fmt="+0.0;-0.0", highlight=True); r += 1
        _cell_style(ws, r, 1, "Total (O/U)", bg=LIGHT_GOLD, border=True)
        _cell_style(ws, r, 2, g["total"], bg=LIGHT_GOLD, align="right",
                    num_fmt="0.0", border=True)
        _cell_style(ws, r, 3, "", bg=LIGHT_GOLD, border=True); r += 2

        # ── ADJUSTMENT METRIC ─────────────────────────────────────────────────
        _section_header(ws, r, 1, "ADJUSTMENT METRIC  (compare to Google Sheets %)", width_cols=3); r += 1
        _row_pair(ws, r, "KenPom Rank", g["debug"]["kenpom_rank_t1"], g["debug"]["kenpom_rank_t2"],
                  t1, t2, num_fmt="0"); r += 1
        _row_pair(ws, r, "NET Rank", g["debug"]["net_rank_t1"], g["debug"]["net_rank_t2"],
                  t1, t2, num_fmt="0"); r += 1
        _row_pair(ws, r, "KP Percentile (0-100)", g["debug"]["kp_pct_t1"], g["debug"]["kp_pct_t2"],
                  t1, t2, num_fmt="0.0%"); r += 1
        _row_pair(ws, r, "NET Percentile (0-100)", g["debug"]["net_pct_t1"], g["debug"]["net_pct_t2"],
                  t1, t2, num_fmt="0.0%"); r += 1
        _row_pair(ws, r, "★ Combined Adj Metric", g["team1_adj_metric"], g["team2_adj_metric"],
                  t1, t2, num_fmt="0.0", highlight=True); r += 2

        # ── PACE ──────────────────────────────────────────────────────────────
        _section_header(ws, r, 1, "PACE", width_cols=3); r += 1
        _row_pair(ws, r, "AdjTempo", g["debug"]["t1_tempo"], g["debug"]["t2_tempo"],
                  t1, t2, num_fmt="0.0"); r += 1
        _cell_style(ws, r, 1, "NCAA Avg Tempo", border=True)
        _cell_style(ws, r, 2, g["debug"]["avg_pace"], align="right", num_fmt="0.0", border=True,
                    bg=LIGHT_BLUE)
        _cell_style(ws, r, 3, "", border=True); r += 1
        _row_pair(ws, r, "★ Projected Pace", g["projected_pace"], g["projected_pace"],
                  t1, t2, num_fmt="0.0", highlight=True); r += 2

        # ── SCORING EFFICIENCY ────────────────────────────────────────────────
        _section_header(ws, r, 1, "SCORING EFFICIENCY", width_cols=3); r += 1
        _row_pair(ws, r, "AdjOE", g["debug"]["t1_adjoe"], g["debug"]["t2_adjoe"], t1, t2, num_fmt="0.0"); r += 1
        _row_pair(ws, r, "AdjDE", g["debug"]["t1_adjde"], g["debug"]["t2_adjde"], t1, t2, num_fmt="0.0"); r += 1
        _row_pair(ws, r, "★ Points Per Possession", g["team1_ppp"], g["team2_ppp"],
                  t1, t2, num_fmt="0.0000", highlight=True); r += 2

        # ── TURNOVERS ─────────────────────────────────────────────────────────
        _section_header(ws, r, 1, "TURNOVERS", width_cols=3); r += 1
        _row_pair(ws, r, "TO_Pct (offense)", g["debug"]["t1_to_pct"], g["debug"]["t2_to_pct"], t1, t2); r += 1
        _row_pair(ws, r, "DTO_Pct (opp defense)", g["debug"]["t2_dto_pct"], g["debug"]["t1_dto_pct"], t1, t2); r += 1
        _cell_style(ws, r, 1, "NCAA Avg TO_Pct", border=True)
        _cell_style(ws, r, 2, g["debug"]["avg_to_pct"], align="right", num_fmt="0.0000", border=True, bg=LIGHT_BLUE)
        _cell_style(ws, r, 3, "", border=True); r += 1
        _row_pair(ws, r, "★ Projected TO (adj)", g["debug"]["t1_to"], g["debug"]["t2_to"],
                  t1, t2, highlight=True); r += 2

        # ── REBOUNDS ──────────────────────────────────────────────────────────
        _section_header(ws, r, 1, "REBOUNDS", width_cols=3); r += 1
        _row_pair(ws, r, "OR_Pct (offense)", g["debug"]["t1_or_pct"], g["debug"]["t2_or_pct"], t1, t2); r += 1
        _row_pair(ws, r, "DOR_Pct (opp allowed)", g["debug"]["t2_dor_pct"], g["debug"]["t1_dor_pct"], t1, t2); r += 1
        _row_pair(ws, r, "★ Projected REB (adj)", g["debug"]["t1_reb"], g["debug"]["t2_reb"],
                  t1, t2, highlight=True); r += 2

        # ── FREE THROWS ───────────────────────────────────────────────────────
        _section_header(ws, r, 1, "FREE THROWS", width_cols=3); r += 1
        _row_pair(ws, r, "FT_Rate (offense)", g["debug"]["t1_ft_rate"], g["debug"]["t2_ft_rate"], t1, t2); r += 1
        _row_pair(ws, r, "DFT_Rate (opp allowed)", g["debug"]["t2_dft_rate"], g["debug"]["t1_dft_rate"], t1, t2); r += 1
        _row_pair(ws, r, "★ Projected FT (adj)", g["debug"]["t1_ft"], g["debug"]["t2_ft"],
                  t1, t2, highlight=True); r += 2

        # ── ADJUSTED POSSESSIONS ──────────────────────────────────────────────
        _section_header(ws, r, 1, "ADJUSTED POSSESSIONS", width_cols=3); r += 1
        _row_pair(ws, r, "★ Adj Possessions", g["debug"]["t1_poss"], g["debug"]["t2_poss"],
                  t1, t2, highlight=True); r += 2

        # ── UNIT SCORE ────────────────────────────────────────────────────────
        _section_header(ws, r, 1, "UNIT SCORE (Height / Experience / Bench)", width_cols=3); r += 1
        _row_pair(ws, r, "AvgHgt", g["debug"]["t1_hgt"], g["debug"]["t2_hgt"], t1, t2, num_fmt="0.0"); r += 1
        _row_pair(ws, r, "Experience (Exp)", g["debug"]["t1_exp"], g["debug"]["t2_exp"], t1, t2, num_fmt="0.00"); r += 1
        _row_pair(ws, r, "Bench", g["debug"]["t1_bench"], g["debug"]["t2_bench"], t1, t2, num_fmt="0.00"); r += 1
        _row_pair(ws, r, "★ Unit Score", g["team1_unit_score"], g["team2_unit_score"],
                  t1, t2, highlight=True); r += 1
        _row_pair(ws, r, "Unit Score Adjustment", g["debug"]["u1_adj"], g["debug"]["u2_adj"],
                  t1, t2, highlight=True); r += 2

        # ── HCA ───────────────────────────────────────────────────────────────
        _section_header(ws, r, 1, "HOME COURT ADVANTAGE", width_cols=3); r += 1
        _row_pair(ws, r, "HCA Applied", g["debug"]["h1_adj"], g["debug"]["h2_adj"],
                  t1, t2, num_fmt="+0.0;-0.0"); r += 2

        # ── VEGAS COMPARISON ──────────────────────────────────────────────────
        _section_header(ws, r, 1, "VEGAS COMPARISON", width_cols=3); r += 1
        _cell_style(ws, r, 1, "Vegas Spread (home)", border=True)
        _cell_style(ws, r, 2, g.get("vegas_spread"), align="right", num_fmt="+0.0;-0.0", border=True, bg=LIGHT_RED)
        _cell_style(ws, r, 3, "", border=True); r += 1
        _cell_style(ws, r, 1, "Vegas Total (O/U)", border=True)
        _cell_style(ws, r, 2, g.get("vegas_total"), align="right", num_fmt="0.0", border=True, bg=LIGHT_RED)
        _cell_style(ws, r, 3, "", border=True); r += 1
        _cell_style(ws, r, 1, "Your Spread vs Vegas", border=True)
        _cell_style(ws, r, 2, g.get("spread"), align="right", num_fmt="+0.0;-0.0", border=True, bg=LIGHT_BLUE)
        _cell_style(ws, r, 3, g.get("vegas_spread"), align="right", num_fmt="+0.0;-0.0", border=True, bg=LIGHT_RED); r += 1
        _cell_style(ws, r, 1, "Spread Difference", border=True)
        _cell_style(ws, r, 2, g.get("spread_edge"), align="right", num_fmt="0.00", border=True); r += 1
        edge = g.get("edge_score")
        edge_bg = "FFD700" if edge and edge > 0.08 else (LIGHT_GREEN if edge and edge > 0.05 else None)
        _cell_style(ws, r, 1, "★ Edge Score (diff/total)", bold=True, border=True)
        _cell_style(ws, r, 2, edge, align="right", num_fmt="0.0000", border=True, bg=edge_bg, bold=True)
        _cell_style(ws, r, 3, "Higher = bigger edge vs Vegas", border=True, font_color="888888"); r += 1
        _cell_style(ws, r, 1, "Source Book", border=True)
        _cell_style(ws, r, 2, g.get("source_book"), border=True); r += 1

    wb.save(filepath)
    print(f"   📊 Debug Excel saved → {filepath}")
    return filepath
